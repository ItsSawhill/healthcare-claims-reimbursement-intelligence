from pathlib import Path
import subprocess
import sys

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))
TABLE_DIR = ROOT / "outputs" / "tables"
REPORT_DIR = ROOT / "outputs" / "reports"
FIGURE_DIR = ROOT / "outputs" / "figures"

from cms_benchmark_loader import apply_cms_or_fallback_benchmarks
from cms_provider_data_loader import (
    CMS_PROVIDER_SOURCE,
    SIMULATED_SOURCE,
    create_cms_provider_service_benchmarks,
    enrich_claims_with_cms_provider_benchmarks,
    load_cms_provider_service_data,
)


def test_required_output_files_exist():
    required_files = [
        TABLE_DIR / "provider_kpis.csv",
        TABLE_DIR / "monthly_trends.csv",
        TABLE_DIR / "reimbursement_benchmarking.csv",
        TABLE_DIR / "utilization_summary.csv",
        TABLE_DIR / "anomalies.csv",
        TABLE_DIR / "forecast_summary.csv",
        TABLE_DIR / "cost_driver_analysis.csv",
        TABLE_DIR / "scenario_rate_change.csv",
        TABLE_DIR / "scenario_utilization_change.csv",
        TABLE_DIR / "scenario_provider_contract_change.csv",
        TABLE_DIR / "scenario_benchmark_alignment.csv",
        TABLE_DIR / "scenario_summary.csv",
        TABLE_DIR / "cms_provider_service_benchmarks.csv",
        REPORT_DIR / "executive_summary.md",
        REPORT_DIR / "executive_workbook.xlsx",
        FIGURE_DIR / "pmpm_trend.png",
        FIGURE_DIR / "provider_efficiency_ranking.png",
        FIGURE_DIR / "utilization_trend_dashboard.png",
        FIGURE_DIR / "scenario_financial_impact.png",
        FIGURE_DIR / "scenario_pmpm_impact.png",
        FIGURE_DIR / "provider_scenario_exposure.png",
        FIGURE_DIR / "benchmark_alignment_impact.png",
    ]
    missing = [str(path) for path in required_files if not path.exists()]
    assert not missing, f"Missing expected outputs: {missing}"


def test_provider_kpi_required_columns():
    provider = pd.read_csv(TABLE_DIR / "provider_kpis.csv")
    required_columns = {
        "provider_id",
        "provider_name",
        "total_claims",
        "total_paid",
        "denial_rate",
        "pmpm_contribution",
        "benchmark_variance_pct",
        "provider_risk_score",
        "provider_risk_tier",
        "provider_segment",
        "high_cost_provider_flag",
    }
    assert required_columns.issubset(provider.columns)
    assert len(provider) > 0
    assert provider["provider_risk_score"].between(0, 100).all()


def test_monthly_trends_required_columns():
    monthly = pd.read_csv(TABLE_DIR / "monthly_trends.csv")
    required_columns = {
        "service_month",
        "total_claims",
        "total_paid",
        "denial_rate",
        "pmpm",
        "claims_per_1000_members",
    }
    assert required_columns.issubset(monthly.columns)
    assert len(monthly) >= 12
    assert (monthly["total_claims"] > 0).all()


def test_reimbursement_and_utilization_outputs_have_required_columns():
    reimbursement = pd.read_csv(TABLE_DIR / "reimbursement_benchmarking.csv")
    utilization = pd.read_csv(TABLE_DIR / "utilization_summary.csv")
    assert {
        "provider_id",
        "service_category",
        "paid_to_billed_rate",
        "allowed_to_billed_rate",
        "benchmark_variance_pct",
        "benchmark_flag",
    }.issubset(reimbursement.columns)
    assert {
        "service_month",
        "service_category",
        "visits_per_1000_members",
        "cost_per_visit",
        "pmpm",
    }.issubset(utilization.columns)


def test_forecast_and_anomaly_outputs_have_required_columns():
    forecast = pd.read_csv(TABLE_DIR / "forecast_summary.csv")
    anomalies = pd.read_csv(TABLE_DIR / "anomalies.csv")
    assert {"forecast_month", "metric", "forecast_value", "method"}.issubset(forecast.columns)
    assert set(forecast["metric"]) == {"total_paid", "total_claims", "pmpm"}
    assert {"anomaly_type", "entity_type", "metric", "metric_value", "severity", "rationale"}.issubset(anomalies.columns)


def test_excel_workbook_sheets():
    workbook = load_workbook(REPORT_DIR / "executive_workbook.xlsx", read_only=True)
    expected_sheets = {
        "Executive Summary",
        "Provider KPIs",
        "Monthly Trends",
        "Reimbursement Benchmarking",
        "Utilization Summary",
        "Anomalies",
        "Forecasts",
        "Scenario Summary",
        "Rate Change Impact",
        "Utilization Impact",
        "Provider Contract Impact",
        "Benchmark Alignment Impact",
    }
    assert expected_sheets.issubset(set(workbook.sheetnames))


def test_scenario_tables_have_required_columns_and_math():
    required_columns = {
        "scenario_name",
        "baseline_paid_amount",
        "simulated_paid_amount",
        "dollar_impact",
        "percent_impact",
        "pmpm_impact",
        "benchmark_variance_impact",
    }
    for filename in [
        "scenario_rate_change.csv",
        "scenario_utilization_change.csv",
        "scenario_provider_contract_change.csv",
        "scenario_benchmark_alignment.csv",
    ]:
        scenario = pd.read_csv(TABLE_DIR / filename)
        assert required_columns.issubset(scenario.columns)
        diff = scenario["simulated_paid_amount"] - scenario["baseline_paid_amount"]
        assert np.allclose(diff, scenario["dollar_impact"], atol=0.01, equal_nan=True)

    summary = pd.read_csv(TABLE_DIR / "scenario_summary.csv")
    assert required_columns.issubset(summary.columns)
    diff = summary["simulated_paid_amount"] - summary["baseline_paid_amount"]
    assert np.allclose(diff, summary["dollar_impact"], atol=0.01, equal_nan=True)


def test_cms_benchmark_loader_fallback_uses_synthetic_source(tmp_path):
    claims = pd.DataFrame(
        {
            "procedure_code": ["99213", "93000"],
            "medicare_benchmark_amount": [100.0, 25.0],
        }
    )
    enriched, source = apply_cms_or_fallback_benchmarks(claims, tmp_path / "missing_cms.csv")
    assert source == "simulated"
    assert set(enriched["benchmark_source"]) == {"simulated"}
    assert enriched["medicare_benchmark_amount"].tolist() == [100.0, 25.0]


def test_cms_provider_loader_missing_file_does_not_fail(tmp_path):
    loaded = load_cms_provider_service_data(tmp_path / "missing_provider_service.csv")
    assert loaded is None


def test_cms_provider_loader_validates_required_columns(tmp_path):
    bad_path = tmp_path / "bad_cms_provider_service.csv"
    pd.DataFrame({"HCPCS_Cd": ["99213"], "Tot_Srvcs": [10]}).to_csv(bad_path, index=False)
    with pytest.raises(ValueError, match="missing required columns"):
        load_cms_provider_service_data(bad_path)


def test_cms_provider_loader_sample_file_creates_benchmark_output(tmp_path):
    cms_path = tmp_path / "cms_provider_service.csv"
    pd.DataFrame(
        {
            "Rndrng_NPI": ["1234567890", "1234567890", "2222222222"],
            "Rndrng_Prvdr_Last_Org_Name": ["Alpha Clinic", "Alpha Clinic", "Beta Group"],
            "Rndrng_Prvdr_State_Abrvtn": ["MD", "MD", "VA"],
            "HCPCS_Cd": ["99213", "99213", "93000"],
            "HCPCS_Desc": ["Office visit", "Office visit", "Electrocardiogram"],
            "Tot_Srvcs": [10, 20, 5],
            "Avg_Sbmtd_Chrg": [150.0, 180.0, 75.0],
            "Avg_Mdcr_Alowd_Amt": [90.0, 100.0, 25.0],
            "Avg_Mdcr_Pymt_Amt": [70.0, 80.0, 20.0],
        }
    ).to_csv(cms_path, index=False)

    loaded = load_cms_provider_service_data(cms_path)
    benchmarks = create_cms_provider_service_benchmarks(loaded)
    output_path = tmp_path / "cms_provider_service_benchmarks.csv"
    benchmarks.to_csv(output_path, index=False)

    assert output_path.exists()
    assert {"procedure_code", "provider_state", "benchmark_level", "avg_medicare_allowed"}.issubset(benchmarks.columns)
    procedure_row = benchmarks[(benchmarks["benchmark_level"] == "procedure_code") & (benchmarks["procedure_code"] == "99213")].iloc[0]
    assert round(procedure_row["avg_medicare_allowed"], 2) == 96.67


def test_synthetic_claims_can_join_to_sample_cms_benchmarks(tmp_path):
    cms = pd.DataFrame(
        {
            "provider_npi": ["123"],
            "provider_name": ["Alpha Clinic"],
            "provider_state": ["MD"],
            "procedure_code": ["99213"],
            "service_description": ["Office visit"],
            "number_of_services": [10],
            "submitted_charge_amount": [150.0],
            "medicare_allowed_amount": [90.0],
            "medicare_payment_amount": [70.0],
            "avg_submitted_charge": [150.0],
            "avg_medicare_allowed": [90.0],
            "avg_medicare_payment": [70.0],
            "medicare_payment_to_charge_ratio": [70.0 / 150.0],
            "allowed_to_charge_ratio": [90.0 / 150.0],
        }
    )
    benchmarks = create_cms_provider_service_benchmarks(cms)
    claims = pd.DataFrame(
        {
            "procedure_code": ["99213", "93000"],
            "allowed_amount": [110.0, 30.0],
            "paid_amount": [85.0, 20.0],
            "medicare_benchmark_amount": [100.0, 25.0],
            "benchmark_source": [SIMULATED_SOURCE, SIMULATED_SOURCE],
        }
    )
    enriched, source = enrich_claims_with_cms_provider_benchmarks(claims, benchmarks)
    assert source == CMS_PROVIDER_SOURCE
    assert enriched.loc[0, "cms_allowed_variance"] == 20.0
    assert enriched.loc[0, "cms_payment_variance"] == 15.0
    assert enriched.loc[0, "benchmark_source"] == CMS_PROVIDER_SOURCE
    assert enriched.loc[1, "benchmark_source"] == SIMULATED_SOURCE


def test_processed_claims_benchmark_source_is_populated():
    claims = pd.read_csv(ROOT / "data" / "processed" / "claims_clean.csv")
    assert "benchmark_source" in claims.columns
    assert claims["benchmark_source"].notna().all()


def test_pipeline_runs_end_to_end():
    result = subprocess.run(
        [sys.executable, "src/run_pipeline.py"],
        cwd=ROOT,
        check=False,
        capture_output=True,
        text=True,
        timeout=60,
    )
    assert result.returncode == 0, result.stderr
    assert "pipeline completed" in result.stdout
